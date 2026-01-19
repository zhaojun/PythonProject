import pandas as pd
import numpy as np
import re
from openpyxl.styles import Font, Alignment, Border, Side
import openpyxl

def extract_bill_by_name(excel_path, target_names, output_folder=".", sheet_name=None):
    """
    按姓名提取账单记录
    :param excel_path: Excel文件路径
    :param target_names: 目标姓名列表（如["沙欧", "王之锐"]）
    :param output_folder: 输出文件夹路径
    :param sheet_name: 工作表名称（None表示读取所有工作表）
    :return: 生成的文件路径列表
    """
    # 读取Excel文件
    if sheet_name:
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
    else:
        df = pd.read_excel(excel_path, header=None)
    
    # 1. 自动识别姓名列（寻找包含"姓名"的列）
    name_col_idx = None
    for col_idx in range(df.shape[1]):
        col_headers = df.iloc[:, col_idx].astype(str).str.contains("姓名", na=False)
        if col_headers.any():
            name_col_idx = col_idx
            break
    
    if name_col_idx is None:
        raise ValueError("未找到'姓名'列，请检查Excel文件表头")
    
    print(f"✅ 已识别姓名列：第{name_col_idx+1}列")
    
    # 2. 按姓名筛选数据行
    result_files = []
    for target_name in target_names:
        # 找到目标姓名所在的行
        name_matches = df.iloc[:, name_col_idx].astype(str).str.contains(target_name, na=False)
        if not name_matches.any():
            print(f"⚠️ 未找到姓名'{target_name}'的数据，跳过")
            continue
        
        # 获取目标行数据（取第一个匹配行，如需多行会匹配可修改）
        target_row = df[name_matches].iloc[0]  # 取第一个匹配行
        print(f"✅ 找到'{target_name}'的数据，开始提取")
        
        # 3. 提取有效数据（去除空白、0、无效文本）
        bill_data = {}
        basic_info_mapping = {
            "编号": None, "姓名": name_col_idx, "地址": None,
            "合计欠款": None, "订餐消费金额": None, "已结": None,
            "2024年12月份消费": None
        }
        
        # 先识别基础信息列位置
        for col_idx in range(df.shape[1]):
            header = str(df.iloc[0, col_idx]) if not pd.isna(df.iloc[0, col_idx]) else ""
            for info_key in basic_info_mapping:
                if info_key in header and basic_info_mapping[info_key] is None:
                    basic_info_mapping[info_key] = col_idx
        
        # 提取基础信息
        for info_key, col_idx in basic_info_mapping.items():
            if col_idx is not None and col_idx < len(target_row):
                value = target_row.iloc[col_idx]
                if not (pd.isna(value) or value == 0 or str(value).strip() in ["0", "没有订餐", "nan"]):
                    bill_data[info_key] = value
        
        # 提取日期消费数据
        date_patterns = [r"(\d+月\d+号)", r"(\d+月\d+日)", r"(\d+月\d+)"]
        for col_idx in range(df.shape[1]):
            if col_idx >= len(target_row):
                continue
            
            header = str(df.iloc[0, col_idx]) if not pd.isna(df.iloc[0, col_idx]) else ""
            value = target_row.iloc[col_idx]
            
            # 跳过无效值
            if pd.isna(value) or value == 0 or str(value).strip() in ["0", "没有订餐", "nan"]:
                continue
            
            # 匹配日期
            for pattern in date_patterns:
                date_match = re.search(pattern, header)
                if date_match:
                    date_str = date_match.group()
                    if "号" not in date_str and "日" not in date_str:
                        date_str += "号"
                    elif "日" in date_str:
                        date_str = date_str.replace("日", "号")
                    
                    # 区分金额和订餐内容
                    if isinstance(value, (int, float)):
                        bill_data[date_str] = value
                    elif "订餐内容" in header:
                        bill_data[f"{date_str}订餐内容"] = value
                    break
        
        # 4. 转为竖版格式并排序
        def sort_key(item):
            key = item[0]
            basic_order = ["编号", "姓名", "地址", "合计欠款", "订餐消费金额", "已结", "2024年12月份消费"]
            if key in basic_order:
                return (0, basic_order.index(key))
            date_match = re.search(r"(\d+)月(\d+)号", key)
            if date_match:
                month = int(date_match.group(1))
                day = int(date_match.group(2))
                year = 2024 if month == 12 else 2025
                return (1, year, month, day, 1 if "订餐内容" in key else 0)
            return (2, key)
        
        sorted_bill = sorted(bill_data.items(), key=sort_key)
        result_df = pd.DataFrame(sorted_bill, columns=["项目", "数值/内容"])
        result_df.insert(0, "序号", range(1, len(result_df) + 1))
        
        # 5. 生成格式化Excel文件
        output_excel = f"{output_folder}/{target_name}_订餐账单.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{target_name}账单"
        
        # 设置样式
        title_font = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
        title_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        data_font = Font(name="微软雅黑", size=10)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        
        # 添加标题
        ws.merge_cells("A1:C1")
        ws["A1"] = f"{target_name}订餐账单（2024年10月-12月）"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        # 添加表头
        headers = ["序号", "项目", "数值/内容"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # 添加数据
        for row_idx, (idx, item, value) in enumerate(result_df.values, 3):
            ws.cell(row=row_idx, column=1, value=idx).font = data_font
            ws.cell(row=row_idx, column=2, value=item).font = data_font
            ws.cell(row=row_idx, column=3, value=value).font = data_font
            for col in [1, 2, 3]:
                ws.cell(row=row_idx, column=col).border = border
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center" if col == 1 else "left", vertical="center")
            ws.row_dimensions[row_idx].height = 25
        
        # 调整列宽
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 55
        
        # 保存文件
        wb.save(output_excel)
        
        # 生成CSV文件
        output_csv = f"{output_folder}/{target_name}_订餐账单.csv"
        result_df.to_csv(output_csv, index=False, encoding="utf-8-sig")
        
        result_files.append((output_excel, output_csv))
        print(f"✅ '{target_name}'的账单已生成：")
        print(f"   - Excel: {output_excel}")
        print(f"   - CSV: {output_csv}")
    
    return result_files

# ------------------- 使用示例 -------------------
if __name__ == "__main__":
    # 配置参数
    EXCEL_FILE_PATH = r"D:\WPS云盘\1214901082\WPS云盘\沈飞账单-最新.xlsx"  # 替换为你的Excel文件路径
    TARGET_NAMES = ["赵勋"]  # 替换为你要提取的姓名列表
    OUTPUT_FOLDER = "."  # 输出文件夹（当前文件夹）
    
    # 执行提取
    try:
        generated_files = extract_bill_by_name(EXCEL_FILE_PATH, TARGET_NAMES, OUTPUT_FOLDER)
        print(f"\n🎉 所有账单提取完成，共生成 {len(generated_files) * 2} 个文件")
    except Exception as e:
        print(f"❌ 提取失败：{str(e)}")